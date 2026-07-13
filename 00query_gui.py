#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# 作者 地图大师 B站：https://space.bilibili.com/41150425 微信公众号：地图大师的漏洞追踪指南 微信号：returnwrong
"""
0.zone API 综合查询工具
========================
基于 0.zone API 的五合一信息查询系统：
  1. 信息系统 (site)      2. 移动端应用 (app)    3. 域名 (domain)
  4. 邮箱 (email)         5. 代码/文档 (code)

运行方式: python 00query_gui.py
"""

import json
import os
import sys
import threading
import tkinter as tk
from tkinter import messagebox, filedialog
from datetime import datetime
from pathlib import Path

import requests
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
from PIL import Image, ImageTk

# ── 全局常量 ────────────────────────────────────────────────────────────────

APP_NAME = "0.zone 综合查询工具"
API_BASE_URL = "https://0.zone/api/data/"
# PyInstaller 打包后，路径指向 exe 所在目录而非临时解压目录
if getattr(sys, 'frozen', False):
    BASE_DIR = Path(sys.executable).parent
else:
    BASE_DIR = Path(__file__).resolve().parent
CONFIG_FILE = BASE_DIR / "00query_config.json"
LOGO_FILE = BASE_DIR / "bate.jpg"
SYNTAX_FILE = BASE_DIR / "语法指南.txt"

# 每页条数选项
PAGE_SIZES = [10, 20, 50, 100]

# 各模块表格列: (字段名, 列标题, 列宽)
COLUMNS_SITE = [
    ("url",       "URL",       220),
    ("title",     "标题",      160),
    ("ip",        "IP",        130),
    ("port",      "端口",       55),
    ("company",   "所属公司",  160),
    ("cms",       "CMS",        80),
    ("os",        "操作系统",   80),
    ("city",      "城市",       70),
    ("province",  "省份",       70),
    ("country",   "国家",       70),
    ("status_code","状态码",    55),
    ("beian",     "备案号",    140),
    ("service",   "服务",       60),
    ("component", "组件",       90),
    ("is_cdn",    "CDN",        45),
    ("timestamp", "更新时间",  140),
]

COLUMNS_APP = [
    ("title",        "应用名称",    180),
    ("type",         "应用类型",    100),
    ("company",      "所属公司",    180),
    ("icp",          "备案号",      150),
    ("app_id",       "AppID/FakeID",180),
    ("introduction", "应用描述",    250),
    ("timestamp",    "入库时间",    140),
    ("timestamp_update","更新时间", 140),
]

COLUMNS_DOMAIN = [
    ("domain",      "域名",       200),
    ("root_domain", "根域名",     160),
    ("company",     "公司",       180),
    ("icp",         "备案号",     150),
    ("toplv_domain","顶级域名",    80),
    ("url",         "子域名",     200),
    ("ip",          "IP",         130),
]

COLUMNS_EMAIL = [
    ("email",       "邮箱",       200),
    ("email_type",  "邮箱类型",    80),
    ("group",       "所属公司",   180),
    ("company",     "公司列表",   180),
    ("mail_domain", "邮箱后缀",   140),
    ("leakage_num", "泄漏次数",    70),
    ("source",      "来源",       200),
    ("timestamp",   "时间",       140),
]

COLUMNS_CODE = [
    ("name",          "文件名",     200),
    ("path",          "路径",       220),
    ("file_extension","类型",        65),
    ("source",        "来源",        90),
    ("owner_login",   "作者",       120),
    ("repo_name",     "仓库名",     160),
    ("score",         "风险值",      60),
    ("timestamp",     "入库时间",   140),
]

COLUMNS_MAP = {
    "site":   COLUMNS_SITE,
    "app":    COLUMNS_APP,
    "domain": COLUMNS_DOMAIN,
    "email":  COLUMNS_EMAIL,
    "code":   COLUMNS_CODE,
}

THEME_LIST = [
    "flatly", "darkly", "superhero", "cosmo", "litera", "minty", "lumen",
    "sandstone", "yeti", "pulse", "united", "journal", "solar", "cyborg",
    "vapor", "simplex", "cerculean",
]

# ── 各模块高级查询字段（字段名, 说明, 可选值, 示例）─────────────────────────
SYNTAX_FIELDS = {
    "site": [
        ("company",            "所属公司",        "任意公司名称",                          "company==网络科技"),
        ("ip",                 "IP / IP段",       "IP 或 CIDR/范围",                       "ip=39.98.171.0/24"),
        ("url",                "URL / 域名",      "完整域名或部分",                        "url==00sec.com"),
        ("title",              "网页标题",        "任意关键词",                            "title==管理后台"),
        ("html_banner",        "HTML 原文",       "页面中出现的文字",                      "html_banner==零零信安"),
        ("banner",             "Banner 信息",     "服务 Banner 内容",                      "banner==nginx"),
        ("res_header",         "响应头",          "HTTP 响应头内容",                       "res_header==nginx"),
        ("port",               "端口",            "80, 443, 22, 21, 8080, 3389, 3306 …",  "port=443"),
        ("component",          "组件",            "nginx, apache, tomcat, iis …",         "component==nginx"),
        ("service",            "服务",            "http, https, ftp, ssh, smtp, mysql …", "service==http"),
        ("tags",               "标签",            "登录页, 后台管理, 防火墙, VPN …",      "tags==登录页"),
        ("os",                 "操作系统",        "windows, linux, ubuntu, centos …",     "os==windows"),
        ("extra_info",         "设备分类",        "路由器, 交换机, 防火墙, 负载均衡 …",  "extra_info==路由器"),
        ("app_name",           "APP 名称",        "任意应用名称",                          "app_name==零零信安"),
        ("status_code",        "HTTP 状态码",     "200, 301, 302, 403, 404, 500, 502 …","status_code=200"),
        ("operator",           "运营商",          "电信, 联通, 移动, 阿里云, 腾讯云 …",  "operator==电信"),
        ("device_type",        "设备类型",        "server, load balancer, firewall …",   "device_type==server"),
        ("versions",           "版本号",          "如 1.0, 2.4.1 …",                     "versions==1.0"),
        ("cms",                "CMS",             "WordPress, DedeCMS, Discuz …",        "cms==WordPress"),
        ("cdn",                "是否 CDN",        "0 (否), 1 (是)",                       "cdn=1"),
        ("icp",                "备案号",          "ICP 备案号关键词",                     "icp==京ICP备"),
        ("icon",               "ICON MD5",        "网站图标的 MD5 哈希值",                "icon=baf2ef6b..."),
        ("ssl_info.issuer_cn", "SSL 颁发者 CN",   "如 Amazon, Let's Encrypt …",          "ssl_info.issuer_cn==Amazon"),
        ("ssl_info.issuer_org","SSL 颁发机构 O",  "如 Amazon, Google …",                 "ssl_info.issuer_org==Amazon"),
        ("ssl_info.subject_cn","SSL 通用名 CN",   "域名或通配符",                         "ssl_info.subject_cn==*.00sec.com"),
        ("ssl_info.subject_org","SSL 关联组织 O", "如 DataDomain, Example Inc …",        "ssl_info.subject_org==DataDomain"),
        ("ssl_info.detail",    "SSL 证书详情",    "证书中出现的任意文字",                 "ssl_info.detail==001"),
        ("lang",               "开发语言",        "java, php, python, go, nodejs …",    "lang==java"),
        ("country",            "国家",            "中国, 美国, 日本, 韩国 …",            "country==中国"),
        ("province",           "省份",            "北京, 上海, 广东, 浙江 …",            "province==广东"),
        ("city",               "城市",            "北京, 上海, 深圳, 杭州 …",            "city==深圳"),
        ("asset_type",         "资产类型",        "ip (IP设备), operating (网站系统)",   "asset_type=operating"),
        ("explore_timestamp",  "录入时间范围",    "week, month, six_month, min_six_month","explore_timestamp=week"),
        ("timestamp",          "更新时间范围",    "week, month, six_month, min_six_month","timestamp=month"),
    ],
    "app": [
        ("company",       "所属公司",       "任意公司名称",                            "company==网络科技"),
        ("name",          "应用名称",       "任意应用名称",                            "name==万能"),
        ("description",   "应用描述",       "描述中的关键词",                          "description==去水印"),
        ("type",          "应用类型",       "微信小程序, 微信公众号, 安卓APK, iOS",    "type=微信小程序"),
        ("icp",           "备案号",         "ICP 备案号",                              "icp=京ICP备"),
        ("domain_list",   "相关域名",       "应用中引用的域名",                        "domain_list==zhushou03.abu.com"),
    ],
    "domain": [
        ("company",       "所属公司",       "任意公司名称",                            "company==网络科技"),
        ("root_domain",   "根域",           "如 00sec.com, baidu.com",                "root_domain==00sec.com"),
        ("ip",            "IP 地址",        "关联的 IP 地址",                          "ip=39.98.171.121"),
        ("domain",        "子域名",         "如 vip.00sec.com, mail.baidu.com",       "domain=vip.00sec.com"),
        ("icp",           "备案号",         "ICP 备案号（建议用 === 逐字匹配）",       "icp===备2022"),
        ("toplv_domain",  "域名后缀",       "com, cn, net, org, gov, edu, io …",     "toplv_domain==cn"),
        ("level",         "域名级别",       "根域名, 子域名",                          "level=子域名"),
    ],
    "email": [
        ("company",       "所属公司",       "任意公司名称",                            "company==网络科技"),
        ("email",         "邮箱地址",       "完整或部分邮箱地址",                      "email==@126.com"),
        ("email_domain",  "邮箱域名",       "如 126.com, qq.com, gmail.com …",       "email_domain==126.com"),
        ("email_type",    "邮箱类型",       "企业邮箱, 个人邮箱",                      "email_type=企业邮箱"),
        ("leakage_num",   "泄露次数",       "数字: 0, 1, 2, 3 …",                    "leakage_num>=5"),
        ("source",        "来源平台",       "baidu.com, skymem.info, github …",       "source=baidu.com"),
        ("msg.tags",      "标签",           "公司名等标签关键词",                      "msg.tags==科技有限公司"),
    ],
    "code": [
        ("name",                       "文件名称",       "文件名关键词",              "name==password"),
        ("related_company",            "关联公司",       "任意公司名称",              "related_company==网络科技"),
        ("code_detail",                "代码/文档原文",  "原文中包含的文字",           "code_detail==password"),
        ("suffix",                     "文件后缀",       "java, py, js, go, html, yml, json, xml, md …","suffix=java"),
        ("repository.description",     "仓库描述",       "GitHub 仓库描述关键词",      "repository.description==爬虫"),
        ("url",                        "来源 URL",       "如 github.com, csdn.net …", "url==github.com"),
        ("detail_parsing.domain_list", "涉及域名",       "代码中出现的域名",           "detail_parsing.domain_list==gmail.com"),
        ("detail_parsing.ip_list",     "涉及 IP",        "代码中出现的 IP",            "detail_parsing.ip_list==10.0"),
        ("detail_parsing.email_list",  "涉及邮箱",       "代码中出现的邮箱",           "detail_parsing.email_list==163.com"),
        ("detail_parsing.telegram_list","涉及 Telegram", "代码中出现的 Telegram 账号",  "detail_parsing.telegram_list==bot"),
        ("source",                     "来源平台",       "github.com, csdn.net, gitee.com …","source=csdn.net"),
        ("tags",                       "标签",           "密码, 配置文件, 数据库, API …","tags==密码"),
        ("type",                       "类型",           "代码, 文档",                  "type=文档"),
        ("keyword",                    "关键词",         "手机号, 密码, 域名, IP …",    "keyword==手机号"),
        ("timestamp",                  "录入时间",       ">= 日期 或 week/month/six_month/min_six_month","timestamp>=2023-10-06"),
        ("timestamp_update",           "更新时间",       ">= 日期 或 week/month/six_month/min_six_month","timestamp_update>=2023-10-06"),
    ],
}

# ── 配置管理 ────────────────────────────────────────────────────────────────

class ConfigManager:
    """API Key 配置文件读写"""

    @staticmethod
    def load() -> dict:
        if CONFIG_FILE.exists():
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    cfg = json.load(f)
                    if isinstance(cfg, dict) and "zone_key_id" in cfg:
                        return cfg
            except Exception:
                pass
        return {"zone_key_id": "", "theme": "flatly"}

    @staticmethod
    def save(config: dict) -> bool:
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=2, ensure_ascii=False)
            return True
        except Exception as e:
            messagebox.showerror("保存失败", f"无法写入配置文件:\n{e}")
            return False

# ── API 客户端 ──────────────────────────────────────────────────────────────

class APIClient:
    """0.zone API 请求封装"""

    def __init__(self, zone_key_id: str):
        self.zone_key_id = zone_key_id
        self.session = requests.Session()
        self.session.headers["Content-Type"] = "application/json"

    def query(self, query: str, query_type: str, page: int = 1,
              pagesize: int = 20, next_key: int = None, **extra) -> dict:
        data: dict = {
            "query": query,
            "query_type": query_type,
            "pagesize": pagesize,
            "zone_key_id": self.zone_key_id,
            **extra,
        }
        if next_key is not None:
            data["next"] = next_key
        else:
            data["page"] = page

        try:
            resp = self.session.post(API_BASE_URL, json=data, timeout=30)
            resp.raise_for_status()
            return resp.json()
        except requests.exceptions.Timeout:
            return {"code": -1, "message": "请求超时，请检查网络连接"}
        except requests.exceptions.ConnectionError:
            return {"code": -1, "message": "无法连接服务器，请检查网络"}
        except Exception as e:
            return {"code": -1, "message": f"请求异常: {str(e)}"}

# ── 详情查看弹窗 ────────────────────────────────────────────────────────────

class DetailViewer(tk.Toplevel):
    """完整 JSON 数据查看窗口"""

    def __init__(self, parent, data: dict, title: str = "数据详情"):
        super().__init__(parent)
        self.title(title)
        self.geometry("820x620")
        self.minsize(420, 320)
        self.transient(parent)

        # 文本框
        tf = ttk.Frame(self)
        tf.pack(fill=BOTH, expand=YES, padx=10, pady=10)

        self.text = tk.Text(tf, wrap=WORD, font=("Consolas", 10),
                           relief=FLAT, borderwidth=0, bg="#f8f9fa")
        sb = ttk.Scrollbar(tf, orient=VERTICAL, command=self.text.yview)
        self.text.configure(yscrollcommand=sb.set)
        self.text.pack(side=LEFT, fill=BOTH, expand=YES)
        sb.pack(side=RIGHT, fill=Y)

        formatted = json.dumps(data, indent=2, ensure_ascii=False)
        self.text.insert("1.0", formatted)
        self.text.config(state=DISABLED)

        # 按钮
        bf = ttk.Frame(self)
        bf.pack(fill=X, padx=10, pady=(0, 10))
        ttk.Button(bf, text="📋 复制全部", command=self._copy,
                   bootstyle="outline-secondary").pack(side=LEFT, padx=3)
        ttk.Button(bf, text="关闭", command=self.destroy,
                   bootstyle="outline-secondary").pack(side=RIGHT, padx=3)

        self._center_on_parent(parent)
        self.grab_set()
        self.focus_force()

    def _copy(self):
        self.clipboard_clear()
        self.clipboard_append(self.text.get("1.0", "end-1c"))
        self.update()

    def _center_on_parent(self, parent):
        self.update_idletasks()
        pw, ph = parent.winfo_width(), parent.winfo_height()
        px, py = parent.winfo_rootx(), parent.winfo_rooty()
        w, h = self.winfo_width(), self.winfo_height()
        x = px + (pw - w) // 2
        y = py + (ph - h) // 2
        self.geometry(f"+{x}+{y}")

# ── 查询面板 (单个模块) ─────────────────────────────────────────────────────

class QueryPanel(ttk.Frame):
    """复用组件：搜索栏 + 结果表格 + 分页器"""

    def __init__(self, parent, query_type: str, get_client, **kw):
        super().__init__(parent, **kw)
        self.query_type = query_type
        self._get_client = get_client
        self._columns = COLUMNS_MAP.get(query_type, [])
        self._data: list = []
        self._total = 0
        self._page = 1
        self._pagesize = 100  # API 最大 pagesize，与网站一致
        self._next_key = None
        self._page_mode = "page"  # "page" or "next"
        self._loading = False

        self._build_ui()
        self._bind_events()

    # ── 构建界面 ────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── 搜索栏 ──
        bar = ttk.Frame(self)
        bar.pack(fill=X, padx=10, pady=(10, 5))

        ttk.Label(bar, text="关键词:", font=("Microsoft YaHei", 10)) \
           .pack(side=LEFT, padx=(0, 6))

        self.query_var = tk.StringVar()
        self.query_entry = ttk.Entry(bar, textvariable=self.query_var, width=36,
                                     font=("Microsoft YaHei", 10))
        self.query_entry.pack(side=LEFT, padx=(0, 10), fill=X, expand=YES)

        # 疑似信息系统勾选框（仅 site）
        if self.query_type == "site":
            self.suspected_var = tk.BooleanVar(value=False)
            ttk.Checkbutton(bar, text="疑似信息系统", variable=self.suspected_var,
                            bootstyle="info-round-toggle") \
               .pack(side=LEFT, padx=(0, 8))

        # 排序下拉（site / app / code）
        if self.query_type in ("site", "app", "code"):
            ttk.Label(bar, text="排序:", font=("Microsoft YaHei", 9)) \
               .pack(side=LEFT, padx=(5, 3))
            self.sort_var = tk.StringVar(value="默认")
            ttk.Combobox(bar, textvariable=self.sort_var,
                         values=["默认", "DESC", "ASC"],
                         width=6, state="readonly") \
               .pack(side=LEFT, padx=(0, 8))

        # 查询语法提示
        hints = {
            "site":   "💡 普通搜索: 公司名/域名  |  高级: ssl_info.detail==example.com",
            "app":    "💡 普通搜索: 公司名/应用名  |  高级: company==公司名 && type==微信公众号  |  icp=备案号 && type==微信小程序",
            "domain": "💡 普通搜索: 公司名/域名    |  高级: root_domain==example.com",
            "email":  "💡 普通搜索: 公司名/邮箱    |  高级: email==@example.com",
            "code":   "💡 普通搜索: 关键词          |  高级: tags==配置文件",
        }
        hint_text = hints.get(self.query_type, "")
        ttk.Label(self, text=hint_text, font=("Microsoft YaHei", 8),
                  foreground="#6c757d").pack(fill=X, padx=10, pady=(0, 2))

        self.search_btn = ttk.Button(bar, text="🔍 查询", command=self._search,
                                     bootstyle="primary")
        self.search_btn.pack(side=LEFT, padx=2)
        ttk.Button(bar, text="清空", command=self._clear,
                   bootstyle="outline-secondary").pack(side=LEFT, padx=2)

        # 进度条（默认隐藏）
        self.progress = ttk.Progressbar(self, mode="indeterminate")

        # ── 状态行 ──
        self.info_var = tk.StringVar(value="就绪 — 请输入关键词开始查询")
        ttk.Label(self, textvariable=self.info_var, font=("Microsoft YaHei", 9),
                  foreground="gray").pack(fill=X, padx=10, pady=(3, 2))

        # ── 表格 ──
        tf = ttk.Frame(self)
        tf.pack(fill=BOTH, expand=YES, padx=10, pady=(0, 5))
        tf.rowconfigure(0, weight=1)
        tf.columnconfigure(0, weight=1)

        col_ids   = [c[0] for c in self._columns]
        col_names = [c[1] for c in self._columns]
        col_widths= [c[2] for c in self._columns]

        self.tv = ttk.Treeview(tf, columns=col_ids, show="headings",
                               bootstyle="primary")
        for cid, cname, cw in zip(col_ids, col_names, col_widths):
            self.tv.heading(cid, text=cname)
            self.tv.column(cid, width=cw, minwidth=40, stretch=True)

        vsb = ttk.Scrollbar(tf, orient=VERTICAL, command=self.tv.yview)
        hsb = ttk.Scrollbar(tf, orient=HORIZONTAL, command=self.tv.xview)
        self.tv.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        self.tv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # ── 分页栏 ──
        pag = ttk.Frame(self)
        pag.pack(fill=X, padx=10, pady=(0, 10))

        ttk.Label(pag, text="每页:", font=("Microsoft YaHei", 9)) \
           .pack(side=LEFT, padx=(0, 4))
        self.psize_var = tk.StringVar(value="100")
        ps_cb = ttk.Combobox(pag, textvariable=self.psize_var,
                             values=[str(v) for v in PAGE_SIZES],
                             width=5, state="readonly")
        ps_cb.pack(side=LEFT, padx=(0, 12))
        ps_cb.bind("<<ComboboxSelected>>", lambda e: self._on_psize_change())

        self.first_btn = ttk.Button(pag, text="⏮ 首页", command=lambda: self._go(1),
                                    bootstyle="outline-secondary")
        self.first_btn.pack(side=LEFT, padx=1)

        self.prev_btn = ttk.Button(pag, text="◀ 上一页",
                                   command=lambda: self._go(self._page - 1),
                                   bootstyle="outline-secondary")
        self.prev_btn.pack(side=LEFT, padx=1)

        ttk.Label(pag, text=" 第", font=("Microsoft YaHei", 9)) \
           .pack(side=LEFT)
        self.page_num_var = tk.StringVar(value="1")
        self.page_entry = ttk.Entry(pag, textvariable=self.page_num_var,
                                    width=5, justify="center",
                                    font=("Microsoft YaHei", 9))
        self.page_entry.pack(side=LEFT, padx=2)
        self.page_entry.bind("<Return>", lambda e: self._go_custom())
        ttk.Label(pag, text="页 ", font=("Microsoft YaHei", 9)) \
           .pack(side=LEFT)

        self.next_btn = ttk.Button(pag, text="下一页 ▶",
                                   command=lambda: self._go(self._page + 1),
                                   bootstyle="outline-secondary")
        self.next_btn.pack(side=LEFT, padx=1)

        self.total_var = tk.StringVar(value="共 0 条")
        ttk.Label(pag, textvariable=self.total_var, font=("Microsoft YaHei", 9),
                  foreground="gray").pack(side=LEFT, padx=(15, 0))

        self.export_btn = ttk.Button(pag, text="📥 导出", command=self._export,
                                     bootstyle="success-outline")
        self.export_btn.pack(side=RIGHT, padx=2)

        self._update_pag_buttons()

    def _bind_events(self):
        self.tv.bind("<Double-1>", self._on_double_click)

    # ── 查询逻辑 ────────────────────────────────────────────────────────

    def _search(self):
        query = self.query_var.get().strip()
        if not query:
            self.info_var.set("⚠ 请输入关键词")
            return
        client = self._get_client()
        if not client:
            self.info_var.set("⚠ 请先在右上角配置 API Key")
            return

        self._page = 1
        self._next_key = None
        self._page_mode = "page"
        self._pagesize = int(self.psize_var.get())
        self._do_fetch(client, query)

    def _do_fetch(self, client: APIClient, query: str, page: int = None,
                  next_key: int = None):
        if self._loading:
            return
        self._loading = True
        self.progress.pack(fill=X, padx=10, pady=(0, 3))
        self.progress.start(8)
        self.search_btn.config(state=DISABLED)
        self.info_var.set("⏳ 正在查询...")

        extra = {}
        if self.query_type == "site" and self.suspected_var.get():
            extra["is_suspected_site"] = 1
        # 排序：仅在用户明确选择 DESC/ASC 时发送，默认不发送（与网站行为一致）
        if self.query_type in ("site", "app", "code"):
            sort_val = self.sort_var.get()
            if sort_val in ("DESC", "ASC"):
                key = "timestamp_sort" if self.query_type == "site" else "timestamp_update_sort"
                extra[key] = sort_val

        def worker():
            try:
                result = client.query(
                    query=query, query_type=self.query_type,
                    page=page or 1, pagesize=self._pagesize,
                    next_key=next_key, **extra)
                self.after(0, lambda: self._handle_result(result))
            except Exception as e:
                self.after(0, lambda: self._handle_error(str(e)))

        threading.Thread(target=worker, daemon=True).start()

    def _handle_result(self, result: dict):
        self._loading = False
        self.progress.stop()
        self.progress.pack_forget()
        self.search_btn.config(state=NORMAL)

        if result.get("code") != 0:
            msg = result.get("message", "查询失败")
            self.info_var.set(f"❌ {msg}")
            self._update_pag_buttons()
            return

        raw = result.get("data", [])

        # 企业黄页疑似信息系统：返回查询语法而非数据列表
        if self.query_type == "site" and self.suspected_var.get() and isinstance(raw, dict):
            self._data = []
            self._total = 0
            self.tv.delete(*self.tv.get_children())
            self.total_var.set("共 0 条")
            self.page_num_var.set("1")
            self._show_suspected_filters(result)
            self.info_var.set("✅ 已生成查询语法，请查看弹窗")
            self._update_pag_buttons()
            return

        self._data = raw if isinstance(raw, list) else []
        self._total = int(result.get("total", 0))

        # API 可能返回 page 或 next 两种分页模式
        resp_page = result.get("page")
        resp_next = result.get("next")

        if resp_page is not None:
            # page 模式：直接使用返回值
            self._page = int(resp_page)
            self._next_key = None
            self._page_mode = "page"
        elif resp_next is not None:
            # next 模式：用 next 做滚动翻页
            self._next_key = resp_next
            self._page_mode = "next"
        else:
            self._page = 1
            self._next_key = None
            self._page_mode = "page"

        # 填充表格
        self.tv.delete(*self.tv.get_children())
        for row in self._data:
            self._insert_row(row)

        # 信息展示
        shown = len(self._data)
        if self._total > shown:
            self.total_var.set(f"显示 {shown} / 共 {self._total} 条")
        else:
            self.total_var.set(f"共 {self._total} 条")

        self.page_num_var.set(str(self._page))
        self._update_pag_buttons()

        usage = result.get("today_api_search_count", {})
        u = usage.get(self.query_type, "")
        info = f"✅ 查询成功 — 显示 {shown} 条，共 {self._total} 条"
        if self._total > shown:
            info += "（可翻页查看更多）"
        if u:
            info += f"  |  今日用量: {u}"
        self.info_var.set(info)

    def _handle_error(self, msg: str):
        self._loading = False
        self.progress.stop()
        self.progress.pack_forget()
        self.search_btn.config(state=NORMAL)
        self.info_var.set(f"❌ {msg}")

    # ── 表格行渲染 ──────────────────────────────────────────────────────

    def _insert_row(self, row: dict):
        qt = self.query_type
        if qt == "site":
            co = row.get("company", "")
            if isinstance(co, list): co = ", ".join(co[:3])
            vals = [row.get("url",""), row.get("title",""), row.get("ip",""),
                    row.get("port",""), co, row.get("cms",""), row.get("os",""),
                    row.get("city",""), row.get("province",""), row.get("country",""),
                    row.get("status_code",""), row.get("beian",""),
                    row.get("service",""), row.get("component",""),
                    "是" if row.get("is_cdn")==1 else "否", row.get("timestamp","")]
        elif qt == "app":
            msg = row.get("msg", {}) or {}
            app_type = row.get("type", "")
            # 根据应用类型显示 AppID 或 FakeID
            if app_type == "微信公众号":
                app_id = msg.get("wechat_fakeid", "") if isinstance(msg, dict) else ""
            elif app_type in ("微信小程序", "安卓APK", "iOS"):
                app_id = msg.get("app_id", "") if isinstance(msg, dict) else ""
            else:
                app_id = (msg.get("app_id", "") or msg.get("wechat_fakeid", "")) if isinstance(msg, dict) else ""
            intro = (msg.get("introduction", "") or "") if isinstance(msg, dict) else ""
            vals = [row.get("title", ""), app_type,
                    row.get("company", "") or row.get("group", ""),
                    row.get("icp", ""), str(app_id),
                    intro[:80],
                    row.get("timestamp", ""), row.get("timestamp_update", "")]
        elif qt == "domain":
            co = row.get("company","")
            if isinstance(co, list): co = ", ".join(co[:3])
            msg = row.get("msg",{}) or {}
            vals = [row.get("domain",""), row.get("root_domain",""), co,
                    row.get("icp",""), row.get("toplv_domain",""),
                    row.get("url",""), msg.get("ip","") if isinstance(msg, dict) else ""]
        elif qt == "email":
            co = row.get("company","")
            if isinstance(co, list): co = ", ".join(co[:3])
            src = row.get("source","")
            if isinstance(src, list): src = ", ".join(src[:5])
            vals = [row.get("email",""), row.get("email_type",""),
                    row.get("group",""), co, str(row.get("mail_domain","")),
                    str(row.get("leakage_num","")), str(src), row.get("timestamp","")]
        elif qt == "code":
            own = row.get("owner",{}) or {}
            rep = row.get("repository",{}) or {}
            vals = [row.get("name",""), row.get("path",""),
                    row.get("file_extension",""), row.get("source",""),
                    own.get("login","") if isinstance(own, dict) else "",
                    rep.get("name","") if isinstance(rep, dict) else "",
                    row.get("score",""), row.get("timestamp","")]
        else:
            vals = []
        self.tv.insert("", END, values=vals)

    # ── 企业黄页查询语法弹窗 ────────────────────────────────────────────

    def _show_suspected_filters(self, result: dict):
        data = result.get("data", {})
        if not isinstance(data, dict):
            return

        items = []
        for key, label in [
            ("company_filter",                "公司名称筛选"),
            ("company_domain_filter",         "域名关联筛选"),
            ("company_tags_filter",           "标签关联筛选"),
            ("company_recored_number_filter", "备案号关联筛选"),
        ]:
            val = data.get(key, "")
            if val:
                items.append((label, val))

        if not items:
            return

        dlg = tk.Toplevel(self)
        dlg.title("企业黄页 — 疑似信息系统查询语法")
        dlg.geometry("700x480")
        dlg.minsize(500, 380)
        dlg.transient(self.winfo_toplevel())

        c = ttk.Frame(dlg, padding=15)
        c.pack(fill=BOTH, expand=YES)

        ttk.Label(c, text="📋 疑似信息系统 — 高级查询语法",
                  font=("Microsoft YaHei", 13, "bold")).pack(anchor=W, pady=(0, 5))
        ttk.Label(c, text="以下查询语法由系统自动生成，可用于进一步搜索关联目标：",
                  font=("Microsoft YaHei", 9), foreground="gray") \
           .pack(anchor=W, pady=(0, 10))

        txt = tk.Text(c, wrap=WORD, font=("Consolas", 10), relief=FLAT,
                      borderwidth=0, padx=8, pady=8, bg="#f8f9fa")
        sb = ttk.Scrollbar(c, orient=VERTICAL, command=txt.yview)
        txt.configure(yscrollcommand=sb.set)
        txt.pack(side=LEFT, fill=BOTH, expand=YES)
        sb.pack(side=RIGHT, fill=Y)

        for label, val in items:
            txt.insert(END, f"【{label}】\n", "bold")
            txt.insert(END, f"  {val}\n\n")
        txt.tag_configure("bold", font=("Consolas", 10, "bold"))
        txt.config(state=DISABLED)

        bf = ttk.Frame(c)
        bf.pack(fill=X, pady=(10, 0))
        for label, val in items:
            ttk.Button(bf, text=f"📋 复制「{label}」",
                       command=lambda v=val: self._copy_to_clipboard(dlg, v),
                       bootstyle="outline-info").pack(side=LEFT, padx=3)
        ttk.Button(bf, text="关闭", command=dlg.destroy,
                   bootstyle="outline-secondary").pack(side=RIGHT)

        # 居中 & 焦点
        dlg.update_idletasks()
        pw = self.winfo_toplevel().winfo_width()
        ph = self.winfo_toplevel().winfo_height()
        px = self.winfo_toplevel().winfo_rootx()
        py = self.winfo_toplevel().winfo_rooty()
        w, h = dlg.winfo_width(), dlg.winfo_height()
        dlg.geometry(f"+{px+(pw-w)//2}+{py+(ph-h)//2}")
        dlg.grab_set()
        dlg.focus_force()

    @staticmethod
    def _copy_to_clipboard(widget, text: str):
        widget.clipboard_clear()
        widget.clipboard_append(text)
        widget.update()

    # ── 分页 ────────────────────────────────────────────────────────────

    def _update_pag_buttons(self):
        if self._page_mode == "next":
            # next 模式：总有"下一页"（除非 total 已知且已到底）
            shown = len(self._data)
            self.first_btn.config(state=DISABLED)
            self.prev_btn.config(state=DISABLED)
            self.next_btn.config(state=NORMAL if shown >= self._pagesize else DISABLED)
        else:
            max_p = max(1, (self._total + self._pagesize - 1) // self._pagesize)
            self.first_btn.config(state=NORMAL if self._page > 1 else DISABLED)
            self.prev_btn.config(state=NORMAL if self._page > 1 else DISABLED)
            self.next_btn.config(state=NORMAL if self._page < max_p else DISABLED)

    def _go(self, target: int):
        if target < 1:
            return
        q = self.query_var.get().strip()
        c = self._get_client()
        if not q or not c:
            return

        if self._page_mode == "next":
            # next 模式：用上次返回的 next_key 翻页
            if target > self._page:
                self._do_fetch(c, q, next_key=self._next_key)
                self._page = target
            else:
                # next 模式不支持回退，重新从第一页搜索
                self._page = 1
                self._next_key = None
                self._do_fetch(c, q, page=1)
        else:
            max_p = max(1, (self._total + self._pagesize - 1) // self._pagesize)
            if target > max_p:
                target = max_p
            self._do_fetch(c, q, page=target)

    def _go_custom(self):
        try:
            self._go(int(self.page_num_var.get()))
        except ValueError:
            self.page_num_var.set(str(self._page))

    def _on_psize_change(self):
        self._pagesize = int(self.psize_var.get())
        self._page = 1
        self._next_key = None
        self._page_mode = "page"
        self._search()

    def _clear(self):
        self.query_var.set("")
        self.tv.delete(*self.tv.get_children())
        self._data.clear()
        self._total = 0
        self._page = 1
        self._next_key = None
        self._page_mode = "page"
        self.info_var.set("就绪 — 请输入关键词开始查询")
        self.total_var.set("共 0 条")
        self.page_num_var.set("1")
        self._update_pag_buttons()

    # ── 导出 ────────────────────────────────────────────────────────────

    def _export(self):
        if not self._data:
            messagebox.showinfo("提示", "当前没有数据可导出", parent=self)
            return
        path = filedialog.asksaveasfilename(
            parent=self,
            defaultextension=".xlsx",
            filetypes=[("Excel 表格", "*.xlsx")],
            initialfile=f"{self.query_type}_{datetime.now():%Y%m%d_%H%M%S}",
        )
        if not path:
            return
        try:
            self._export_xlsx(path)
            self.info_var.set(f"✅ 已导出: {path}")
        except ImportError:
            try:
                csv_path = path.replace(".xlsx", ".csv")
                self._export_csv(csv_path)
                self.info_var.set(f"✅ 已导出(CSV): {csv_path}")
            except Exception as e2:
                messagebox.showerror("导出失败", str(e2), parent=self)
        except Exception as e:
            messagebox.showerror("导出失败", str(e), parent=self)

    def _get_display_row(self, row: dict) -> dict:
        """提取与表格列一致的展示数据，返回 {列key: 展示值}"""
        qt = self.query_type
        result = {}
        if qt == "site":
            co = row.get("company", "")
            if isinstance(co, list): co = ", ".join(co[:3])
            result = {
                "url": row.get("url",""), "title": row.get("title",""),
                "ip": row.get("ip",""), "port": row.get("port",""),
                "company": str(co), "cms": row.get("cms",""),
                "os": row.get("os",""), "city": row.get("city",""),
                "province": row.get("province",""), "country": row.get("country",""),
                "status_code": row.get("status_code",""), "beian": row.get("beian",""),
                "service": row.get("service",""), "component": row.get("component",""),
                "is_cdn": "是" if row.get("is_cdn")==1 else "否",
                "timestamp": row.get("timestamp",""),
            }
        elif qt == "app":
            msg = row.get("msg", {}) or {}
            app_type = row.get("type", "")
            if app_type == "微信公众号":
                app_id = msg.get("wechat_fakeid", "") if isinstance(msg, dict) else ""
            elif app_type in ("微信小程序", "安卓APK", "iOS"):
                app_id = msg.get("app_id", "") if isinstance(msg, dict) else ""
            else:
                app_id = (msg.get("app_id", "") or msg.get("wechat_fakeid", "")) if isinstance(msg, dict) else ""
            intro = (msg.get("introduction", "") or "") if isinstance(msg, dict) else ""
            result = {
                "title": row.get("title",""), "type": app_type,
                "company": row.get("company","") or row.get("group",""),
                "icp": row.get("icp",""), "app_id": str(app_id),
                "introduction": intro[:80],
                "timestamp": row.get("timestamp",""),
                "timestamp_update": row.get("timestamp_update",""),
            }
        elif qt == "domain":
            co = row.get("company","")
            if isinstance(co, list): co = ", ".join(co[:3])
            msg = row.get("msg",{}) or {}
            result = {
                "domain": row.get("domain",""), "root_domain": row.get("root_domain",""),
                "company": str(co), "icp": row.get("icp",""),
                "toplv_domain": row.get("toplv_domain",""),
                "url": row.get("url",""),
                "ip": msg.get("ip","") if isinstance(msg, dict) else "",
            }
        elif qt == "email":
            co = row.get("company","")
            if isinstance(co, list): co = ", ".join(co[:3])
            src = row.get("source","")
            if isinstance(src, list): src = ", ".join(src[:5])
            result = {
                "email": row.get("email",""), "email_type": row.get("email_type",""),
                "group": row.get("group",""), "company": str(co),
                "mail_domain": str(row.get("mail_domain","")),
                "leakage_num": str(row.get("leakage_num","")),
                "source": str(src), "timestamp": row.get("timestamp",""),
            }
        elif qt == "code":
            own = row.get("owner",{}) or {}
            rep = row.get("repository",{}) or {}
            result = {
                "name": row.get("name",""), "path": row.get("path",""),
                "file_extension": row.get("file_extension",""),
                "source": row.get("source",""),
                "owner_login": own.get("login","") if isinstance(own, dict) else "",
                "repo_name": rep.get("name","") if isinstance(rep, dict) else "",
                "score": row.get("score",""), "timestamp": row.get("timestamp",""),
            }
        return result

    def _export_xlsx(self, path: str):
        """导出为 XLSX — 列与表格展示一致，首列为序号"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = self.query_type

        # 列定义：序号 + 表格展示列
        col_keys = ["序号"] + [c[0] for c in self._columns]
        col_names = ["序号"] + [c[1] for c in self._columns]

        # 表头样式
        hdr_font = Font(name="Microsoft YaHei", bold=True, color="FFFFFF", size=10)
        hdr_fill = PatternFill(start_color="2471A3", end_color="2471A3", fill_type="solid")
        hdr_align = Alignment(horizontal="center", vertical="center")
        seq_align = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )
        cell_font = Font(name="Microsoft YaHei", size=9)
        cell_align = Alignment(vertical="center", wrap_text=True)

        # 写表头
        for col_idx, name in enumerate(col_names, 1):
            cell = ws.cell(row=1, column=col_idx, value=name)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = hdr_align
            cell.border = thin_border

        # 写数据行
        for row_idx, row in enumerate(self._data, 2):
            display = self._get_display_row(row)
            # 序号列
            seq_cell = ws.cell(row=row_idx, column=1, value=row_idx - 1)
            seq_cell.font = cell_font
            seq_cell.alignment = seq_align
            seq_cell.border = thin_border
            # 数据列
            for col_idx, key in enumerate(col_keys[1:], 2):
                val = display.get(key, "")
                cell = ws.cell(row=row_idx, column=col_idx, value=str(val) if val else "")
                cell.font = cell_font
                cell.alignment = cell_align
                cell.border = thin_border

        # 自动列宽
        for col_idx in range(1, len(col_keys) + 1):
            max_len = len(str(col_names[col_idx - 1])) * 2
            for row_idx in range(2, min(len(self._data) + 2, 102)):
                cv = ws.cell(row=row_idx, column=col_idx).value
                if cv:
                    max_len = max(max_len, len(str(cv)))
            from openpyxl.utils import get_column_letter
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 55)

        ws.freeze_panes = "B2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(path)

    def _export_csv(self, path: str):
        """回退 CSV 导出 — 列与表格展示一致"""
        import csv
        col_keys = [c[0] for c in self._columns]
        col_names = ["序号"] + [c[1] for c in self._columns]
        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            w = csv.writer(f)
            w.writerow(col_names)
            for i, row in enumerate(self._data, 1):
                display = self._get_display_row(row)
                vals = [i] + [display.get(k, "") for k in col_keys]
                w.writerow(vals)

    # ── 双击详情 ────────────────────────────────────────────────────────

    def _on_double_click(self, event):
        sel = self.tv.selection()
        if not sel:
            return
        idx = self.tv.index(sel[0])
        if idx < len(self._data):
            DetailViewer(self, self._data[idx], f"数据详情 — 第 {idx+1} 条")

# ── API Key 配置对话框 ──────────────────────────────────────────────────────

class KeyConfigDialog(tk.Toplevel):
    """API Key 配置窗口"""

    def __init__(self, parent, config: dict, on_save_callback):
        super().__init__(parent)
        self.title("API Key 配置")
        # ★ 大幅增加默认尺寸，确保所有内容+按钮可见
        self.geometry("600x460")
        self.minsize(520, 420)
        self.transient(parent)

        self._config = config
        self._on_save = on_save_callback
        self._saved = False

        self._build_ui()
        self._center_on_parent(parent)
        self.after(100, self._do_grab)

    def _do_grab(self):
        self.grab_set()
        self.key_entry.focus_set()

    def _center_on_parent(self, parent):
        self.update_idletasks()
        pw = parent.winfo_width()
        ph = parent.winfo_height()
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        w = self.winfo_reqwidth()
        h = self.winfo_reqheight()
        self.geometry(f"+{px + (pw - w) // 2}+{py + (ph - h) // 2}")

    def _build_ui(self):
        # 外层容器 fill entire window
        outer = ttk.Frame(self)
        outer.pack(fill=BOTH, expand=YES)

        # 使用 grid 布局，更可控
        outer.columnconfigure(0, weight=1)
        outer.columnconfigure(1, weight=0)  # for show checkbox
        # 给最后一行（按钮行上方）留出额外空间
        outer.rowconfigure(8, weight=1)

        row = 0
        # ── 标题 ──
        ttk.Label(outer, text="⚙ API Key 配置",
                  font=("Microsoft YaHei", 15, "bold")) \
            .grid(row=row, column=0, columnspan=2, sticky="w",
                  padx=32, pady=(24, 6))
        row += 1
        ttk.Label(outer, text="请输入您的 0.zone API Key，用于调用数据查询接口。",
                  font=("Microsoft YaHei", 9), foreground="gray") \
            .grid(row=row, column=0, columnspan=2, sticky="w",
                  padx=32, pady=(0, 20))
        row += 1

        # ── API Key 标签 ──
        ttk.Label(outer, text="API Key:",
                  font=("Microsoft YaHei", 10, "bold")) \
            .grid(row=row, column=0, columnspan=2, sticky="w",
                  padx=32, pady=(0, 8))
        row += 1

        # ── API Key 输入 + 显示勾选框 ──
        self.key_var = tk.StringVar(value=self._config.get("zone_key_id", ""))
        self.key_entry = tk.Entry(
            outer, textvariable=self.key_var, show="●",
            font=("Consolas", 11), relief=SOLID, borderwidth=1,
        )
        self.key_entry.grid(row=row, column=0, sticky="ew",
                            padx=(32, 8), pady=(0, 4), ipady=3)

        self.show_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(outer, text="👁 显示", variable=self.show_var,
                        command=self._toggle_show, bootstyle="info-round-toggle") \
            .grid(row=row, column=1, sticky="w", padx=(0, 32))
        row += 1

        # ── 操作提示 ──
        ttk.Label(outer,
                  text="💡 输入 Key 后按 Enter 键，或点击下方「保存配置」按钮",
                  font=("Microsoft YaHei", 8), foreground="#6c757d") \
            .grid(row=row, column=0, columnspan=2, sticky="w",
                  padx=32, pady=(4, 20))
        row += 1

        # ── 分隔线 ──
        ttk.Separator(outer).grid(row=row, column=0, columnspan=2,
                                  sticky="ew", padx=32, pady=(0, 16))
        row += 1

        # ── 主题选择 ──
        ttk.Label(outer, text="界面主题:",
                  font=("Microsoft YaHei", 10, "bold")) \
            .grid(row=row, column=0, columnspan=2, sticky="w",
                  padx=32, pady=(0, 8))
        row += 1

        self.theme_var = tk.StringVar(value=self._config.get("theme", "flatly"))
        ttk.Combobox(outer, textvariable=self.theme_var, values=THEME_LIST,
                     width=24, state="readonly") \
            .grid(row=row, column=0, columnspan=2, sticky="ew",
                  padx=32, pady=(0, 20))
        row += 1

        # ── 按钮 ── (用 row weight 推到底部)
        bf = ttk.Frame(outer)
        bf.grid(row=row, column=0, columnspan=2, sticky="ew",
                padx=32, pady=(0, 24))
        bf.columnconfigure(0, weight=1)
        ttk.Button(bf, text="💾 保存配置 (Enter)", command=self._save,
                   bootstyle="success").grid(row=0, column=1, padx=4)
        ttk.Button(bf, text="取消 (Esc)", command=self.destroy,
                   bootstyle="outline-secondary").grid(row=0, column=2)

        # ── 键盘绑定 ──
        self.key_entry.bind("<Return>", lambda e: self._save())
        self.bind("<Escape>", lambda e: self.destroy())

    def _toggle_show(self):
        self.key_entry.config(show="" if self.show_var.get() else "●")
        self.key_entry.focus_set()

    def _save(self):
        key = self.key_var.get().strip()
        if not key:
            messagebox.showwarning("提示", "API Key 不能为空，请输入有效的 Key。",
                                   parent=self)
            self.key_entry.focus_set()
            return

        self._config["zone_key_id"] = key
        self._config["theme"] = self.theme_var.get()

        if not ConfigManager.save(self._config):
            return

        self._saved = True
        self._on_save(self._config)

        self.grab_release()
        masked = f"{key[:8]}****{key[-4:]}" if len(key) > 12 else key
        messagebox.showinfo("✅ 保存成功",
                            f"API Key 已保存！\n\n"
                            f"Key: {masked}\n"
                            f"主题: {self._config['theme']}\n\n"
                            f"可以开始查询了。",
                            parent=self)
        self.destroy()

# ── 主窗口 ──────────────────────────────────────────────────────────────────

# ── 语法指南窗口 ────────────────────────────────────────────────────────────

class SyntaxGuideWindow(tk.Toplevel):
    """查询语法参考指南 — 结构化展示"""

    def __init__(self, parent):
        super().__init__(parent)
        self.title("查询语法指南")
        self.geometry("960x680")
        self.minsize(700, 500)
        self.transient(parent)

        self._build_ui()
        self._center_on_parent(parent)
        self.after(50, lambda: (self.grab_set(), self.focus_force()))

    def _center_on_parent(self, parent):
        self.update_idletasks()
        pw = parent.winfo_width()
        ph = parent.winfo_height()
        px = parent.winfo_rootx()
        py = parent.winfo_rooty()
        w = self.winfo_width()
        h = self.winfo_height()
        self.geometry(f"+{px + (pw - w) // 2}+{py + (ph - h) // 2}")

    def _build_ui(self):
        outer = ttk.Frame(self, padding=(10, 10))
        outer.pack(fill=BOTH, expand=YES)

        ttk.Label(outer, text="📖 0.zone 查询语法参考指南",
                  font=("Microsoft YaHei", 15, "bold")).pack(anchor=W, pady=(0, 2))
        ttk.Label(outer, text="双击任意参数行可复制查询语法到剪贴板",
                  font=("Microsoft YaHei", 9), foreground="gray") \
           .pack(anchor=W, pady=(0, 8))

        # 选项卡：基础 + 5个模块
        nb = ttk.Notebook(outer, bootstyle="primary")
        nb.pack(fill=BOTH, expand=YES)

        # Tab 1: 查询基础
        base_tab = ttk.Frame(nb)
        nb.add(base_tab, text="📋 查询基础")
        self._build_base_tab(base_tab)

        # Tabs 2-6: 各模块参数
        modules = [
            ("🏢 信息系统",   "site"),
            ("📱 移动端应用", "app"),
            ("🌐 域名",      "domain"),
            ("📧 邮箱",      "email"),
            ("💻 代码/文档",  "code"),
        ]
        from collections import OrderedDict
        for mod_name, mod_key in modules:
            tab = ttk.Frame(nb)
            nb.add(tab, text=mod_name)
            self._build_module_tab(tab, mod_key)

        # 底部
        bf = ttk.Frame(outer)
        bf.pack(fill=X, pady=(8, 0))
        ttk.Label(bf, text="💡 提示: 双击任意行可复制语法  |  Esc 关闭",
                  font=("Microsoft YaHei", 8), foreground="gray").pack(side=LEFT)
        ttk.Button(bf, text="关闭 (Esc)", command=self.destroy,
                   bootstyle="outline-secondary").pack(side=RIGHT)
        self.bind("<Escape>", lambda e: self.destroy())

    def _build_base_tab(self, parent):
        """查询基础: 书写规范 + 逻辑连接符 + 条件匹配符"""
        canvas = tk.Canvas(parent, highlightthickness=0)
        sb = ttk.Scrollbar(parent, orient=VERTICAL, command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw", tags="inner")
        canvas.configure(yscrollcommand=sb.set)

        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)
        canvas.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")
        # 鼠标滚轮
        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<MouseWheel>", _on_mousewheel)
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        # ── 书写规范 ──
        ttk.Label(inner, text="📝 查询条件书写规范",
                  font=("Microsoft YaHei", 12, "bold")).pack(anchor=W, pady=(10, 5))
        rules = [
            "• 条件的值如果包含特殊字符，必须用双引号 \"\" 包裹，如: html_banner===\"id=\"__next\"\"",
            "• 建议使用双引号包裹查询值，并采用逐字符匹配（===）进行精确查询",
            "• 双引号内不能包含未转义的逻辑运算符 || 或 && 作为值的一部分",
        ]
        for r in rules:
            ttk.Label(inner, text=r, font=("Microsoft YaHei", 9),
                      foreground="#333", wraplength=860, justify=LEFT) \
               .pack(anchor=W, padx=10, pady=1)

        # ── 逻辑连接符 ──
        ttk.Label(inner, text="🔗 逻辑连接符",
                  font=("Microsoft YaHei", 12, "bold")).pack(anchor=W, pady=(15, 5))
        logic_data = [
            ("&&", "且 (AND)", "满足A并且满足B", "company==保险 && email===163.com"),
            ("||", "或 (OR)",  "满足A或者满足B", "company==电力 || company==能源"),
            ("( )", "分组",   "用括号确定优先级", "(company==学院 || company==学校) && (name==李)"),
        ]
        self._make_table(inner, ["符号", "含义", "说明", "示例"], logic_data)

        # ── 条件匹配符 ──
        ttk.Label(inner, text="🔍 条件匹配符",
                  font=("Microsoft YaHei", 12, "bold")).pack(anchor=W, pady=(15, 5))
        match_data = [
            ("=",   "完全相等",  "company=北京零零信安科技有限公司", "company=\"\" 查询为空"),
            ("==",  "包含",     "company==microsoft  title==健康码", "包含某单词或词语"),
            ("===", "逐字符匹配","email===edu.com.cn", "内容不是完整单词时使用"),
            ("=^",  "开始为",   "email=^noreply", "匹配字段开头"),
            ("=$",  "结尾为",   "email=$sina.com.cn", "匹配字段结尾"),
            ("=!",  "不等于",   "company=!\"\"", "company==通信 && email=!edu.com.cn"),
            ("==!", "不包含",   "company==信息 || title==!湖南", "排除包含某词的结果"),
            ("=!^", "开始不为", "title==科技 && title=!^灰科技", "排除以某词开头的"),
            ("=!$", "结尾不为", "title==科技 && title=!$科技处", "排除以某词结尾的"),
            (">",   "大于",     "timestamp>2023-01-01", "用于时间判断"),
            (">=",  "大于等于", "timestamp>=2023-01-01", "用于时间判断"),
            ("<",   "小于",     "timestamp<2023-12-31", "用于时间判断"),
            ("<=",  "小于等于", "timestamp<=2023-12-31", "用于时间判断"),
        ]
        self._make_table(inner, ["符号", "含义", "示例", "备注"], match_data)

    def _build_module_tab(self, parent, mod_key: str):
        """构建单个模块的参数表格"""
        canvas = tk.Canvas(parent, highlightthickness=0)
        sb = ttk.Scrollbar(parent, orient=VERTICAL, command=canvas.yview)
        inner = ttk.Frame(canvas)
        inner.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.create_window((0, 0), window=inner, anchor="nw", tags="inner")
        canvas.configure(yscrollcommand=sb.set)

        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)
        canvas.grid(row=0, column=0, sticky="nsew")
        sb.grid(row=0, column=1, sticky="ns")

        def _on_mw(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
        canvas.bind("<MouseWheel>", _on_mw)
        canvas.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mw))
        canvas.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        ttk.Label(inner, text=f"🔍 {self._mod_label(mod_key)} — 高级查询参数",
                  font=("Microsoft YaHei", 12, "bold")).pack(anchor=W, pady=(10, 8))

        params = SYNTAX_FIELDS.get(mod_key, [])
        if params:
            rows = []
            for field, desc, values, example in params:
                rows.append((field, desc, values, example or f"{field}==关键词"))
            self._make_table(inner, ["字段名", "说明", "可选值", "示例"], rows)
        else:
            ttk.Label(inner, text="(暂无参数数据)", font=("Microsoft YaHei", 9)) \
               .pack(anchor=W, padx=10)

    def _mod_label(self, key: str) -> str:
        return {"site": "信息系统", "app": "移动端应用", "domain": "域名",
                "email": "邮箱", "code": "代码/文档"}.get(key, key)

    def _make_table(self, parent, headers: list, rows: list):
        """创建带样式的 Treeview 表格，双击行复制"""
        tf = ttk.Frame(parent)
        tf.pack(fill=X, padx=10, pady=(0, 12))

        cols = [f"c{i}" for i in range(len(headers))]
        tv = ttk.Treeview(tf, columns=cols, show="headings",
                          bootstyle="light", height=min(len(rows), 20))
        for col, hdr in zip(cols, headers):
            tv.heading(col, text=hdr)
        # 4列模式：字段名 / 说明 / 可选值 / 示例
        if len(headers) == 4:
            tv.column(cols[0], width=170, minwidth=120)  # 字段名
            tv.column(cols[1], width=110, minwidth=80)   # 说明
            tv.column(cols[2], width=280, minwidth=180)  # 可选值
            tv.column(cols[3], width=280, minwidth=180)  # 示例
        else:
            tv.column(cols[0], width=150, minwidth=100)
            for col in cols[1:]:
                tv.column(col, width=340, minwidth=150)

        vsb = ttk.Scrollbar(tf, orient=VERTICAL, command=tv.yview)
        tv.configure(yscrollcommand=vsb.set)
        tv.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        tf.rowconfigure(0, weight=1)
        tf.columnconfigure(0, weight=1)

        for row in rows:
            tv.insert("", END, values=list(row))

        # 双击行复制
        def on_double_click(event):
            sel = tv.selection()
            if sel:
                vals = tv.item(sel[0], "values")
                if vals:
                    field = vals[0]
                    text = f"{field}==" if field else ""
                    self.clipboard_clear()
                    self.clipboard_append(text)
                    self.update()
        tv.bind("<Double-1>", on_double_click)

# ── 主窗口 ──────────────────────────────────────────────────────────────────

class MainApp(ttk.Window):
    """应用程序主窗口"""

    def __init__(self):
        self.config = ConfigManager.load()
        theme = self.config.get("theme", "flatly")
        super().__init__(themename=theme)

        self.title(APP_NAME)
        self.geometry("1280x800")
        self.minsize(960, 620)

        self._client: APIClient | None = None
        if self.config.get("zone_key_id"):
            self._client = APIClient(self.config["zone_key_id"])

        self.panels: dict[str, QueryPanel] = {}

        self._build_ui()
        self._center_on_screen()

        # 绑定快捷键
        self.bind("<Control-f>", lambda e: self._focus_search())
        self.bind("<Control-k>", lambda e: self._open_config())

    # ── 布局 ────────────────────────────────────────────────────────────

    def _build_ui(self):
        # ── 顶部栏 ──
        hdr = ttk.Frame(self, padding=(15, 12))
        hdr.pack(fill=X)

        # 左侧：Logo 图片
        tf = ttk.Frame(hdr)
        tf.pack(side=LEFT)

        # 加载 bate.jpg 作为 Logo
        self._logo_img = None  # 保持引用防止 GC
        if LOGO_FILE.exists():
            try:
                img = Image.open(LOGO_FILE).resize((38, 38), Image.LANCZOS)
                self._logo_img = ImageTk.PhotoImage(img)
                ttk.Label(tf, image=self._logo_img).pack(side=LEFT, padx=(0, 10))
            except Exception:
                ttk.Label(tf, text="🔎", font=("Segoe UI Emoji", 26)).pack(side=LEFT, padx=(0, 10))
        else:
            ttk.Label(tf, text="🔎", font=("Segoe UI Emoji", 26)).pack(side=LEFT, padx=(0, 10))

        tt = ttk.Frame(tf)
        tt.pack(side=LEFT)
        ttk.Label(tt, text=APP_NAME, font=("Microsoft YaHei", 17, "bold")).pack(anchor=W)
        ttk.Label(tt, text="基于 0.zone API · 企业资产信息综合查询平台 作者：地图大师和&AI",
                  font=("Microsoft YaHei", 9), foreground="gray").pack(anchor=W)

        # 右侧：Key 状态 + 语法指南 + 配置按钮
        kf = ttk.Frame(hdr)
        kf.pack(side=RIGHT)
        self.key_status_var = tk.StringVar()
        self._refresh_key_label()
        ttk.Label(kf, textvariable=self.key_status_var,
                  font=("Microsoft YaHei", 9)).pack(side=LEFT, padx=(0, 10))
        ttk.Button(kf, text="📖 语法指南", command=self._open_syntax_guide,
                   bootstyle="outline-info").pack(side=LEFT, padx=3)
        ttk.Button(kf, text="⚙ 配置 Key", command=self._open_config,
                   bootstyle="outline-warning").pack(side=LEFT)

        ttk.Separator(self).pack(fill=X)

        # ── 选项卡 ──
        self.notebook = ttk.Notebook(self, bootstyle="primary")
        self.notebook.pack(fill=BOTH, expand=YES, padx=10, pady=(6, 8))

        tabs = [
            ("🏢  信息系统",    "site"),
            ("📱  移动端应用",  "app"),
            ("🌐  域名",       "domain"),
            ("📧  邮箱",       "email"),
            ("💻  代码/文档",   "code"),
        ]
        for name, qtype in tabs:
            panel = QueryPanel(self.notebook, qtype, self._get_client)
            self.notebook.add(panel, text=name)
            self.panels[qtype] = panel

        # ── 底部状态栏 ──
        ttk.Separator(self).pack(fill=X)
        ft = ttk.Frame(self, padding=(15, 8))
        ft.pack(fill=X)
        ttk.Label(ft, text="© 2026 0.zone 查询工具  |  API: 0.zone/api/data  |  作者: 地图大师",
                  font=("Microsoft YaHei", 8), foreground="gray").pack(side=LEFT)
        self.status_var = tk.StringVar(value="就绪")
        ttk.Label(ft, textvariable=self.status_var,
                  font=("Microsoft YaHei", 8), foreground="gray").pack(side=RIGHT)

    def _center_on_screen(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw = self.winfo_screenwidth()
        sh = self.winfo_screenheight()
        self.geometry(f"+{(sw - w) // 2}+{(sh - h) // 2}")

    # ── Key 管理 ────────────────────────────────────────────────────────

    def _get_client(self) -> APIClient | None:
        if self._client and self._client.zone_key_id:
            return self._client
        return None

    def _refresh_key_label(self):
        k = self.config.get("zone_key_id", "")
        if k:
            self.key_status_var.set("🟢 Key 已配置")
        else:
            self.key_status_var.set("🔴 未配置 API Key")

    def _open_config(self):
        KeyConfigDialog(self, self.config, self._on_config_saved)

    def _open_syntax_guide(self):
        SyntaxGuideWindow(self)

    def _on_config_saved(self, config: dict):
        self.config = config
        self._client = APIClient(config["zone_key_id"])
        self._refresh_key_label()
        self.status_var.set("✅ API Key 已保存，可以开始查询")

        # 如果需要切换主题则提示
        new_theme = config.get("theme", "flatly")
        if new_theme != self.style.theme.name.lower():
            if messagebox.askyesno("主题变更",
                                   f"主题已改为「{new_theme}」\n\n需要重启应用才能生效。\n是否立即退出？",
                                   parent=self):
                self.destroy()

    def _focus_search(self):
        idx = self.notebook.index(self.notebook.select())
        panel = list(self.panels.values())[idx]
        panel.query_entry.focus_set()

# ── 入口 ────────────────────────────────────────────────────────────────────

def main():
    # 检查依赖
    missing = []
    for mod_name, pkg_name in [("ttkbootstrap", "ttkbootstrap"), ("requests", "requests")]:
        try:
            __import__(mod_name)
        except ImportError:
            missing.append(pkg_name)

    if missing:
        print("正在安装必要依赖...")
        import subprocess
        for pkg in missing:
            subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "-q"])
        print("安装完成，请重新运行程序。")
        sys.exit(0)

    app = MainApp()
    app.mainloop()

if __name__ == "__main__":
    main()
